import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

class ExcelManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.ensure_file_exists()

    def ensure_file_exists(self):
        if not os.path.exists(self.file_path):
            print(f"Création du fichier {self.file_path} avec données de test...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Rapports"
            
            # NOUVEAU : Séparation de 'nom' et 'prenom'
            headers = ["nom", "prenom", "professionnel", "numero_commande", "status", "date_envoi", "date_reception"]
            ws.append(headers)
            
            ws.append(["PHAN", "Raphaël", "Dr. Goetz", "N/A", "Envoyé", "19/06/2026", ""])
            ws.append(["CIBIEL", "Naomie", "Dr. Goetz", "Z0130819440", "Reçu", "19/06/2026", "20/06/2026"])
            
            wb.save(self.file_path)

    def read_data(self):
        wb = load_workbook(self.file_path)
        ws = wb.active
        data = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]: # Vérifie que nom et prénom existent
                data.append({
                    "nom": row[0],
                    "prenom": row[1],
                    "professionnel": row[2],
                    "numero_commande": row[3],
                    "status": row[4],
                    "date_envoi": row[5],
                    "date_reception": row[6]
                })
        return data

    def update_status_to_recu(self, nom, prenom):
        wb = load_workbook(self.file_path)
        ws = wb.active
        date_jour = datetime.now().strftime("%d/%m/%Y")
        
        for row in ws.iter_rows(min_row=2):
            # On vérifie maintenant le nom ET le prénom
            if row[0].value == nom and row[1].value == prenom:
                row[4].value = "Reçu"      # Index 4 = status
                row[6].value = date_jour   # Index 6 = date_reception
                break
        wb.save(self.file_path)

    def add_new_report(self, nom, prenom, numero_commande):
        wb = load_workbook(self.file_path)
        ws = wb.active
        
        date_jour = datetime.now().strftime("%d/%m/%Y")
        if not numero_commande:
            numero_commande = "N/A"
            
        nouvelle_ligne = [nom, prenom, "Dr. Goetz", numero_commande, "Envoyé", date_jour, ""]
        ws.append(nouvelle_ligne)
        wb.save(self.file_path)